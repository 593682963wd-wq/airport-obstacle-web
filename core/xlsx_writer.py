"""生成含公式的 XLSX 工作簿 — 复刻样例Excel全部结构与公式"""
from __future__ import annotations
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from core.models import Airport, Runway, QFU, Obstacle
from templates.constants import M_TO_FT


# ── 样式 ──────────────────────────────────
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 细则sheet名(固定)
XIZE_SHEET_NAME = "细则（来源AD 2.10机场障碍物）"


def generate_xlsx(airport: Airport, filepath: str,
                  qfu_headings: dict[str, float] | None = None):
    """
    生成XLSX工作簿.
    qfu_headings: {qfu_ident: 磁方位} 的映射(从AIP解析获得)
    """
    wb = Workbook()

    # 1) 跑道数据
    ws_rwy = wb.active
    ws_rwy.title = "跑道数据"
    _build_runway_data_sheet(ws_rwy, airport, qfu_headings or {})

    # 2) 细则
    ws_xize = wb.create_sheet(XIZE_SHEET_NAME)
    _build_xize_sheet(ws_xize, airport)

    # 3) 每个离场方向 + 偏角页
    num_obs = len(airport.obstacles)
    rwy_data_row_map = _get_rwy_data_row_map(airport)
    # D3 always uses the reference runway's length
    ref_idx = airport.reference_runway_idx
    if 0 <= ref_idx < len(airport.runways):
        main_rwy_length = airport.runways[ref_idx].max_length
    else:
        main_rwy_length = airport.runways[0].max_length if airport.runways else 0

    for rwy_idx, rwy in enumerate(airport.runways):
        for qfu in rwy.main_qfus:
            heading = _get_heading(qfu, rwy, qfu_headings or {})
            dep_elev = _get_departure_end_elevation(qfu, rwy)
            # 主方向sheet (F6=0)
            ws_main = wb.create_sheet(qfu.ident)
            _build_direction_sheet(
                ws_main, airport, rwy, qfu, heading, dep_elev,
                num_obs, turn_angle=0.0, is_turn_sheet=False,
                main_rwy_length=main_rwy_length,
            )

            # 偏角页
            turn = qfu.departure_turn_angle
            if turn < 0:
                sheet_name = f"{qfu.ident} 左偏{abs(int(turn))}°"
            elif turn > 0:
                sheet_name = f"{qfu.ident} 右偏{int(turn)}°"
            else:
                sheet_name = f"{qfu.ident} 偏0°"
            ws_turn = wb.create_sheet(sheet_name)
            _build_direction_sheet(
                ws_turn, airport, rwy, qfu, heading, dep_elev,
                num_obs, turn_angle=turn, is_turn_sheet=True,
                main_rwy_length=main_rwy_length,
            )

    # 4) 全局格式: 所有单元格水平垂直居中, 列宽自适应
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ws in wb.worksheets:
        _apply_center_and_autofit(ws, center_align)

    wb.save(filepath)
    return wb


def _apply_center_and_autofit(ws, center_align: Alignment):
    """对 sheet 应用水平+垂直居中、强制自动换行、列宽自适应、行高自适应。

    UI 优化要点 (v1.1.0):
    - 所有单元格强制 wrap_text=True, 避免长文本被截断;
    - 列宽上限提升到 60 (中文字符按 2 计);
    - 估算多行后, 行高按 max(15 × 行数, 18) 自动放大, 防止文字被遮挡。
    """
    col_max_width: dict[int, float] = {}
    row_max_lines: dict[int, int] = {}

    for row in ws.iter_rows():
        for cell in row:
            # 强制水平+垂直居中, 强制自动换行
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            if cell.value is None:
                continue
            val_str = str(cell.value)
            # 公式给一个合理默认宽度即可
            if val_str.startswith("="):
                char_len = 12
                line_count = 1
            else:
                lines = val_str.split("\n")
                line_count = len(lines)
                # 列宽按单行最长计算, 中文字符按 2 计
                char_len = 0
                for line in lines:
                    line_len = sum(2 if ord(c) > 127 else 1 for c in line)
                    char_len = max(char_len, line_len)
            col = cell.column
            col_max_width[col] = max(col_max_width.get(col, 0), char_len)
            r = cell.row
            row_max_lines[r] = max(row_max_lines.get(r, 1), line_count)

    # 列宽: 加 padding, 最小 10, 最大 60 (放大上限避免折断关键列)
    for col_idx, width in col_max_width.items():
        col_letter = get_column_letter(col_idx)
        adjusted = min(max(width + 4, 10), 60)
        ws.column_dimensions[col_letter].width = adjusted

    # 行高: 按单元格里实际折行数估算 (\n 明文换行 + 超列宽自动折行)
    for r, base_lines in row_max_lines.items():
        wrapped = base_lines
        for c_idx, col_w in col_max_width.items():
            cell = ws.cell(row=r, column=c_idx)
            if cell.value is None:
                continue
            text = str(cell.value)
            if text.startswith("="):
                continue
            # 实际列宽 (上限 60)
            actual_w = min(max(col_w + 4, 10), 60)
            for line in text.split("\n"):
                line_w = sum(2 if ord(c) > 127 else 1 for c in line)
                if line_w > actual_w:
                    wrapped = max(wrapped, base_lines + (line_w // actual_w))
        ws.row_dimensions[r].height = max(18, wrapped * 16)


# ── 跑道数据页 ──────────────────────────────

def _build_runway_data_sheet(ws, airport: Airport, qfu_headings: dict):
    ws["A1"] = "制作人"
    ws["C1"] = "制作日期"
    ws["A2"] = "校对人"
    ws["C2"] = "校对日期"
    ws["B3"] = "跑道物理特性"
    ws["B3"].font = HEADER_FONT

    # 表头行 4 (从B列开始, 匹配金标准)
    headers = ["机场代码", "跑道号", "磁方位", "离地端高", "TORA", "净空道", "停止道"]
    for col_idx, h in enumerate(headers, 2):  # start at col B
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    # 数据行: 只列主方向QFU, 从第5行开始
    data_row = 5
    for rwy in airport.runways:
        for qfu in rwy.main_qfus:
            heading = _get_heading(qfu, rwy, qfu_headings)
            dep_elev = _get_departure_end_elevation(qfu, rwy)
            ws.cell(row=data_row, column=2, value=airport.icao)
            ws.cell(row=data_row, column=3, value=qfu.ident)
            ws.cell(row=data_row, column=4, value=heading)
            ws.cell(row=data_row, column=5, value=dep_elev)
            ws.cell(row=data_row, column=6, value=qfu.tora)
            ws.cell(row=data_row, column=7, value=qfu.clearway)
            ws.cell(row=data_row, column=8, value=qfu.stopway)
            for col in range(2, 9):
                ws.cell(row=data_row, column=col).border = THIN_BORDER
            data_row += 1

    # 汇总表格区块 — 每个主方向QFU列出 K="是" 的障碍物
    # 留空行
    data_row += 1
    summary_start = data_row

    # 障碍物索引映射(用于定位方向sheet行号)
    obs_id_map = {id(obs): i for i, obs in enumerate(airport.obstacles)}
    FIRST_OBS_ROW = 11  # 方向sheet中障碍物起始行

    # 按参考跑道生成标签
    ref_idx = airport.reference_runway_idx
    if 0 <= ref_idx < len(airport.runways):
        ref_rwy = airport.runways[ref_idx]
        ref_qfus = ref_rwy.main_qfus
        if len(ref_qfus) >= 2:
            ref_label = f"{ref_qfus[0].ident}/{ref_qfus[1].ident}号跑道中心"
        else:
            ref_label = "跑道中心"
    else:
        ref_label = "跑道中心"

    ws.cell(row=data_row, column=1, value="汇总表格")
    ws.cell(row=data_row, column=2, value=ref_label).font = HEADER_FONT
    data_row += 1

    for rwy in airport.runways:
        for qfu in rwy.main_qfus:
            sheet_name = qfu.ident
            sn = f"'{sheet_name}'"

            # 汇总表头
            sum_headers = ["跑道号", "距离跑道末端距离（米）", "相对跑道末端高（米）",
                           "相对跑道末端高（英尺）", "相对坡度",
                           "无梯度要求\n复飞：标准梯度"]
            for col_idx, h in enumerate(sum_headers, 1):
                c = ws.cell(row=data_row, column=col_idx, value=h)
                c.font = HEADER_FONT
                c.border = THIN_BORDER
                c.alignment = Alignment(wrap_text=True)
            data_row += 1

            # 筛选 K="是" 的障碍物, 按距离排序
            obs_list = sorted(
                [r for r in qfu.obstacle_results if r.is_obstacle],
                key=lambda r: r.dist_from_end
            )

            if obs_list:
                first_row = True
                for obs_r in obs_list:
                    if first_row:
                        ws.cell(row=data_row, column=1, value=qfu.ident)
                        first_row = False

                    # 计算方向sheet行号
                    obs_idx = obs_id_map.get(id(obs_r.obstacle))
                    if obs_idx is not None:
                        dir_row = FIRST_OBS_ROW + obs_idx
                        # 使用公式引用方向sheet(自动更新)
                        ws.cell(row=data_row, column=2).value = f"={sn}!S{dir_row}"
                        ws.cell(row=data_row, column=3).value = f"={sn}!T{dir_row}"
                        ws.cell(row=data_row, column=4).value = \
                            f"=ROUND({sn}!T{dir_row}*{M_TO_FT},0)"
                        ws.cell(row=data_row, column=5).value = \
                            f"=IF({sn}!S{dir_row}>0,{sn}!T{dir_row}/{sn}!S{dir_row},\"\")"
                    else:
                        # Fallback: 使用计算值
                        ws.cell(row=data_row, column=2, value=obs_r.dist_from_end)
                        ws.cell(row=data_row, column=3, value=obs_r.ht_above_end)
                        ws.cell(row=data_row, column=4,
                                value=round(obs_r.ht_above_end * M_TO_FT))
                        if obs_r.dist_from_end > 0:
                            ws.cell(row=data_row, column=5,
                                    value=obs_r.ht_above_end / obs_r.dist_from_end)

                    # F: 标注(序号 + A标签)
                    ws.cell(row=data_row, column=6, value=obs_r.comment_label)
                    # G: 被遮蔽标记
                    if obs_r.is_shielded:
                        ws.cell(row=data_row, column=7, value="被遮蔽")

                    for col in range(1, 8):
                        ws.cell(row=data_row, column=col).border = THIN_BORDER
                    data_row += 1
            else:
                # 无障碍物计算结果时, 用IF公式为每个障碍物创建条件行
                num_obs = len(airport.obstacles)
                ws.cell(row=data_row, column=1, value=qfu.ident)
                for i in range(num_obs):
                    dir_row = FIRST_OBS_ROW + i
                    ws.cell(row=data_row, column=2).value = \
                        f'=IF({sn}!K{dir_row}="是",{sn}!S{dir_row},"")'
                    ws.cell(row=data_row, column=3).value = \
                        f'=IF({sn}!K{dir_row}="是",{sn}!T{dir_row},"")'
                    ws.cell(row=data_row, column=4).value = \
                        f'=IF({sn}!K{dir_row}="是",ROUND({sn}!T{dir_row}*{M_TO_FT},0),"")'
                    ws.cell(row=data_row, column=5).value = \
                        f'=IF(AND({sn}!K{dir_row}="是",{sn}!S{dir_row}>0),' \
                        f'{sn}!T{dir_row}/{sn}!S{dir_row},"")'
                    ws.cell(row=data_row, column=6).value = \
                        f'=IF({sn}!K{dir_row}="是",{sn}!B{dir_row},"")'
                    for col in range(1, 8):
                        ws.cell(row=data_row, column=col).border = THIN_BORDER
                    data_row += 1

            data_row += 1  # 方向之间留空行


# ── 细则页 ──────────────────────────────────

def _build_xize_sheet(ws, airport: Airport):
    ws["A1"] = "本表禁止删行操作，否则影响之后表格的数据调用"
    ws["A1"].font = Font(color="FF0000", bold=True)

    headers = ["序号", "障碍物", "磁方位", "距离", "坐标", "海拔高度",
               "控制障碍物及涉及航段/起飞航径区重要障碍物", "备注"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.border = THIN_BORDER

    for i, obs in enumerate(airport.obstacles):
        row = i + 2
        ws.cell(row=row, column=1, value=obs.seq)
        ws.cell(row=row, column=2, value=obs.name)
        ws.cell(row=row, column=3, value=obs.bearing)
        ws.cell(row=row, column=4, value=obs.distance)
        ws.cell(row=row, column=5, value=obs.coordinate)
        ws.cell(row=row, column=6, value=obs.elevation_m)
        ws.cell(row=row, column=7, value=obs.remark_control)
        ws.cell(row=row, column=8, value=obs.note)
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN_BORDER


# ── 方向分析页 ──────────────────────────────

def _build_direction_sheet(
    ws, airport: Airport, runway: Runway, qfu: QFU,
    heading: float, dep_elev: float, num_obs: int,
    turn_angle: float, is_turn_sheet: bool,
    main_rwy_length: int = 0,
):
    """构建方向分析sheet (主方向或偏角页)"""
    xize = f"'{XIZE_SHEET_NAME}'"
    rwy_len = main_rwy_length or runway.max_length

    # ── 参数区(行1-7) ──
    ws["C1"] = "基本参数"
    ws["C1"].font = HEADER_FONT
    ws["I1"] = "扇区边界"
    ws["I1"].font = HEADER_FONT

    # 行2
    ws["C2"] = "主跑道信息"
    ws["E2"] = "所分析跑道信息"
    ws["F2"] = "交叉跑道信息"
    ws["G2"] = "机场基准点相对跑道中心位移"
    ws["I2"] = "=D3/2+D6"
    ws["J2"] = 90

    # 行3
    ws["C3"] = "长度（m）"
    ws["D3"] = rwy_len
    ws["E3"] = "离地端x轴位移"
    ws["E4"] = qfu.departure_x_offset  # E4: x轴位移
    ws["F3"] = "x轴正方向旋转角度"
    ws["F4"] = qfu.rotation_angle      # F4: 旋转角度
    ws["G3"] = "（沿x轴正方向）"
    ws["G4"] = qfu.arp_offset          # G4: ARP偏移
    ws["I3"] = "=I2"
    ws["J3"] = -90

    # 行4
    ws["C4"] = "磁方向（度）"
    # D4 存参考方位(heading - F4), 使公式 H=C-D4-F4 = C-heading
    ws["D4"] = heading - qfu.rotation_angle

    ws["I4"] = "=6480+$I$3"
    ws["J4"] = 900

    # 行5
    ws["C5"] = "所分析跑道离地点标高（m）"
    ws["D5"] = dep_elev
    ws["E5"] = "跑道中心沿y轴位移"
    ws["E6"] = qfu.lateral_offset      # E6: 横向偏移
    ws["F5"] = "离场转弯相对于跑道离场磁方向(度)"
    ws["I5"] = "=I4"
    ws["J5"] = -900

    # 行6 — F6 是离场转弯角, 核心人工输入位
    ws["C6"] = "所分析跑道净空道（m）"
    ws["D6"] = qfu.clearway
    ws["F6"] = turn_angle
    ws["F6"].fill = YELLOW_FILL
    ws["F6"].comment = Comment(
        "请从离场图读取此离场方向的平均转弯角。\n左偏填负值、右偏填正值、直飞填0。",
        "系统提示"
    )
    ws["G6"] = "离场左转为\"-\"，右转为\"+\""

    ws["I6"] = "=10000+I5"
    ws["J6"] = 900
    ws["S6"] = "\"跑道数据\"表格所需"

    # 行7
    ws["I7"] = "=I6"
    ws["J7"] = -900

    # ── 障碍物分析区(行9起) ──
    # 行9: 列标题
    # 行9: 区域标题(匹配金标准)
    section_headers_9 = {
        "B": "障碍物信息",
        "L": "输入值",
        "N": "坐标转换",
        "S": "PEP输入值",
    }
    for col_letter, title in section_headers_9.items():
        cell = ws[f"{col_letter}9"]
        cell.value = title
        cell.font = HEADER_FONT

    # 行10: 列标题
    col_headers = {
        "B": "序号",
        "C": "磁方位（度）",
        "D": "距离（m）",
        "E": "海拔高度（m）",
        "F": "1.2%梯度面高度（m）",
        "G": "是否穿过",
        "H": "方位差",
        "I": "X",
        "J": "Y",
        "K": "是否为障碍物",
        "L": "DIST",
        "M": "HT",
        "N": "角度(弧度)",
        "O": "x",
        "P": "y",
        "Q": "包线",
        "R": "保护区",
        "S": "距离跑道末端距离（米）",
        "T": "相对跑道末端高（米）",
    }
    for col_letter, title in col_headers.items():
        cell = ws[f"{col_letter}10"]
        cell.value = title
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True)

    # 行10: 副标题(可选)
    # 行11起: 每行对应细则表第2行起的障碍物

    FIRST_DATA_ROW = 11
    for i in range(num_obs):
        r = FIRST_DATA_ROW + i         # Excel行号
        xize_row = i + 2               # 细则表行号(第1行是表头, 数据从第2行)

        # B: 序号 = 细则!A{xize_row}
        ws[f"B{r}"] = f"={xize}!A{xize_row}"

        # C: 磁方位 = 细则!C{xize_row}
        ws[f"C{r}"] = f"={xize}!C{xize_row}"

        # D: 距离 = 细则!D{xize_row}
        ws[f"D{r}"] = f"={xize}!D{xize_row}"

        # E: 海拔高度 = 细则!F{xize_row}
        ws[f"E{r}"] = f"={xize}!F{xize_row}"

        # F: 1.2%梯度面高度 = $D$5+(O{r}-$D$6)*0.012
        ws[f"F{r}"] = f"=$D$5+(O{r}-$D$6)*0.012"

        # G: 是否穿过 = IF(AND(E{r}>=F{r},S{r}>0),"是","否")
        ws[f"G{r}"] = f'=IF(AND(E{r}>=F{r},S{r}>0),"是","否")'

        # H: 方位差 = C{r}-$D$4-$F$4
        ws[f"H{r}"] = f"=C{r}-$D$4-$F$4"

        # I: X = D{r}*COS(H{r}*PI()/180)-$E$4+$G$4-$D$3/2
        ws[f"I{r}"] = f"=D{r}*COS(H{r}*PI()/180)-$E$4+$G$4-$D$3/2"

        # J: Y = D{r}*SIN(H{r}*PI()/180)+$E$6
        ws[f"J{r}"] = f"=D{r}*SIN(H{r}*PI()/180)+$E$6"

        # K: 是否为障碍物 = IF(AND(G{r}="是",R{r}="是"),"是","否")
        ws[f"K{r}"] = f'=IF(AND(G{r}="是",R{r}="是"),"是","否")'

        # L: DIST = INT(O{r})
        ws[f"L{r}"] = f"=INT(O{r})"

        # M: HT = E{r}-D$5
        ws[f"M{r}"] = f"=E{r}-D$5"

        # N: 角度(弧度) = ABS(ATAN2(I{r},J{r})-F$6*PI()/180)
        ws[f"N{r}"] = f"=ABS(ATAN2(I{r},J{r})-F$6*PI()/180)"

        # O: x(沿离场路径) = (I{r}^2+J{r}^2)^0.5*COS(N{r})
        ws[f"O{r}"] = f"=(I{r}^2+J{r}^2)^0.5*COS(N{r})"

        # P: y(垂直离场路径) = ABS((I{r}^2+J{r}^2)^0.5*SIN(N{r}))
        ws[f"P{r}"] = f"=ABS((I{r}^2+J{r}^2)^0.5*SIN(N{r}))"

        # Q: 包线 = IF((ABS(O{r})-$D$6)>6480,900,90+0.125*(ABS(O{r})-$D$6))
        ws[f"Q{r}"] = f"=IF((ABS(O{r})-$D$6)>6480,900,90+0.125*(ABS(O{r})-$D$6))"

        # R: 保护区 = IF(AND(ABS(P{r})<ABS(Q{r}),O{r}>0),"是","否")
        ws[f"R{r}"] = f'=IF(AND(ABS(P{r})<ABS(Q{r}),O{r}>0),"是","否")'

        # S: 距离跑道末端距离 = L{r}
        ws[f"S{r}"] = f"=L{r}"

        # T: 相对跑道末端高 = M{r}
        ws[f"T{r}"] = f"=M{r}"


# ── 辅助 ──────────────────────────────────

def _get_rwy_data_row_map(airport: Airport) -> dict[str, int]:
    """获取每个QFU在跑道数据页中的行号"""
    row_map = {}
    data_row = 5
    for rwy in airport.runways:
        for qfu in rwy.qfus:
            row_map[qfu.ident] = data_row
            data_row += 1
    return row_map


def _guess_heading(qfu: QFU) -> float:
    """从ident粗估磁方位(后续由PDF解析覆盖)"""
    ident = qfu.ident.split()[0] if " " in qfu.ident else qfu.ident
    num_str = ""
    for ch in ident:
        if ch.isdigit():
            num_str += ch
        else:
            break
    if num_str:
        return int(num_str) * 10
    return 0


def _get_heading(qfu: QFU, rwy: Runway, qfu_headings: dict) -> float:
    """获取QFU磁方位: 优先用qfu_headings > qfu.magnetic_heading > 猜测"""
    if qfu.ident in qfu_headings:
        return qfu_headings[qfu.ident]
    if qfu.magnetic_heading:
        return qfu.magnetic_heading
    return _guess_heading(qfu)


def _get_departure_end_elevation(qfu: QFU, rwy: Runway) -> float:
    """离地端标高 = 对端QFU的入口标高(飞机离地点在跑道远端)"""
    main_qfus = rwy.main_qfus
    if len(main_qfus) >= 2:
        if qfu.ident == main_qfus[0].ident:
            opp = main_qfus[1]
        elif qfu.ident == main_qfus[1].ident:
            opp = main_qfus[0]
        else:
            opp = None
        if opp and opp.threshold_elevation > 0:
            return opp.threshold_elevation
    # fallback: 自身标高
    return qfu.threshold_elevation
